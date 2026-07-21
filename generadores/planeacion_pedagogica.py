"""Genera la Planeación Pedagógica (formato oficial GFPI-F-134) del SENA.

Estrategia: usa la plantilla oficial GFPI-F-134.xlsx como base y solo llena las celdas
de datos, preservando encabezado institucional, celdas fusionadas y formato oficial.

Estructura de datos:
{
  "fecha_elaboracion": "2026-07-21",
  "programa": "Técnico en Integración de Operaciones Logísticas",
  "modalidad": "Presencial",
  "codigo_programa": "137136 - Versión 1",
  "proyecto_formativo": "REGISTRAR...",
  "codigo_proyecto": "PF-2026-001",
  "equipo_curricular": "Carlos Barrios",
  "regional_centro": "Regional Guajira - Centro Industrial y de Energías Alternativas",

  "filas": [
    {
      "fase": "Planear",
      "actividad_proyecto": "Identificar los principios...",
      "competencia": "220201501 - Aplicar conocimientos...",
      "raps": "1. Aplicación...\n2. Organizar...",
      "saberes_conceptos": "Fuerza, masa, peso, fricción...",
      "saberes_proceso": "Identificar principios físicos...",
      "criterios_evaluacion": "Identifica los principios físicos...",
      "actividades_aprendizaje": "Guía de aprendizaje S1 RA-01: Leyes de Newton...",
      "horas_directas": 48,
      "horas_independientes": 48,
      "descripcion_evidencia": "Guía resuelta, video experimental, propuesta...",
      "estrategias_didacticas": "ABP, simulación PhET, exposición dialogada...",
      "ambiente": "Aula de sistemas con conexión a internet",
      "materiales": "Computadores, video beam, simuladores PhET",
      "instructores": "Carlos Barrios",
      "observaciones": "",
    }
  ]
}
"""
import shutil
from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "GFPI-F-134.xlsx"

# Mapa de columnas en la tabla (fila 18 en adelante)
COLS_TABLA = {
    "fase":                    1,   # A
    "actividad_proyecto":      2,   # B
    "competencia":             3,   # C
    "raps":                    4,   # D
    "saberes_conceptos":       5,   # E
    "saberes_proceso":         6,   # F
    "criterios_evaluacion":    7,   # G
    "actividades_aprendizaje": 8,   # H
    "horas_directas":          9,   # I
    "horas_independientes":    10,  # J
    "descripcion_evidencia":   11,  # K
    "estrategias_didacticas":  12,  # L
    "ambiente":                13,  # M
    "materiales":              14,  # N
    "instructores":            15,  # O
    "observaciones":           16,  # P
}

FILA_INICIO_TABLA = 18
COLOR_ALT = "F0F0F0"


def _obtener_estilo_referencia(ws, fila_ref=18):
    """Copia el estilo de la fila de referencia (18) para reutilizarlo en filas nuevas."""
    estilos = {}
    for col_idx in range(1, 17):
        cell = ws.cell(row=fila_ref, column=col_idx)
        estilos[col_idx] = {
            "font": copy(cell.font),
            "alignment": copy(cell.alignment),
            "border": copy(cell.border),
            "fill": copy(cell.fill),
        }
    return estilos


def _aplicar_estilo(cell, estilo, fila_par=False):
    """Aplica el estilo copiado a la celda, con alternado gris."""
    cell.font = estilo["font"]
    cell.alignment = Alignment(
        horizontal=estilo["alignment"].horizontal or "left",
        vertical="top",
        wrap_text=True,
    )
    cell.border = estilo["border"]
    if fila_par:
        cell.fill = PatternFill(start_color=COLOR_ALT, end_color=COLOR_ALT, fill_type="solid")


def generar_planeacion(datos: dict, ruta_salida: str) -> str:
    """Genera el archivo de planeación pedagógica llenando la plantilla oficial."""
    shutil.copy(TEMPLATE_PATH, ruta_salida)
    wb = load_workbook(ruta_salida)
    ws = wb["FASE"]

    # ===== CABECERA (columna E porque las celdas están fusionadas E-P) =====
    ws["E9"] = datos.get("fecha_elaboracion", "")
    ws["E10"] = datos.get("programa", "")
    ws["E11"] = datos.get("modalidad", "Presencial")
    ws["E12"] = datos.get("codigo_programa", "")
    ws["E13"] = datos.get("proyecto_formativo", "")
    ws["E14"] = datos.get("codigo_proyecto", "")
    ws["E15"] = datos.get("equipo_curricular", "")
    ws["K15"] = datos.get("regional_centro", "")

    # Formato de las celdas de cabecera
    for coord in ["E9", "E10", "E11", "E12", "E13", "E14", "E15", "K15"]:
        celda = ws[coord]
        celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        celda.font = Font(name="Calibri", size=10, color="2C2C2C")

    # ===== FILAS DE LA TABLA =====
    filas = datos.get("filas", [])
    if filas:
        # Copiar estilo de la fila 18 (que es la de referencia en la plantilla)
        estilos_ref = _obtener_estilo_referencia(ws, fila_ref=18)

        # Ajustar altura mínima de las filas de datos para que se lea bien el wrap
        for i, fila_datos in enumerate(filas):
            num_fila = FILA_INICIO_TABLA + i
            fila_par = (i % 2 == 1)

            for campo, col_idx in COLS_TABLA.items():
                cell = ws.cell(row=num_fila, column=col_idx)
                valor = fila_datos.get(campo, "")
                if campo in ("horas_directas", "horas_independientes"):
                    try:
                        cell.value = int(valor) if valor not in ("", None) else 0
                    except (ValueError, TypeError):
                        cell.value = valor
                else:
                    cell.value = str(valor) if valor is not None else ""
                _aplicar_estilo(cell, estilos_ref[col_idx], fila_par=fila_par)

            # Altura mínima
            ws.row_dimensions[num_fila].height = 120

    # Anchuras razonables para columnas
    anchuras = {
        1: 12, 2: 25, 3: 22, 4: 30, 5: 22, 6: 22, 7: 26, 8: 26,
        9: 10, 10: 10, 11: 24, 12: 22, 13: 18, 14: 22, 15: 20, 16: 18,
    }
    for col, ancho in anchuras.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho

    wb.save(ruta_salida)
    return ruta_salida
