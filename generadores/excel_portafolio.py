"""Genera el Excel consolidado del portafolio del instructor.

Estructura:
  - Hoja 1: "Resumen" — lista de todos los aprendices con estado general
  - Una hoja por aprendiz — con su plan de trabajo detallado, editable
"""
import io
import re
from datetime import date
from pathlib import Path
from typing import List, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DARK = "2C2C2C"
LIGHT = "F0F0F0"
MID = "D0D0D0"
WHITE = "FFFFFF"


def _borde_gris():
    lado = Side(border_style="thin", color=MID)
    return Border(top=lado, bottom=lado, left=lado, right=lado)


def _sanitizar_nombre_hoja(nombre: str) -> str:
    """Excel limita nombre de hoja a 31 chars, sin ciertos caracteres."""
    limpio = re.sub(r'[\[\]\*/\\?:]', '', str(nombre))
    return limpio[:31] or "Sin nombre"


def generar_excel_portafolio(
    aprendices_planes: List[Dict],
    datos_guia: Dict,
    instructor: Dict,
    ruta_salida: str,
) -> str:
    """
    aprendices_planes: lista de dicts con:
      - datos_aprendiz: {nombre, apellidos, correo, ficha, programa}
      - cronograma: lista devuelta por calcular_cronograma()
      - archivo_pdf: str (path del PDF generado)
      - correo_enviado: bool
      - fecha_envio: str
    datos_guia: {programa, competencia, proyecto_formativo, fase}
    instructor: {nombre, cargo}
    """
    wb = Workbook()

    # ===== Hoja RESUMEN =====
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    _construir_resumen(ws_resumen, aprendices_planes, datos_guia, instructor)

    # ===== Una hoja por aprendiz =====
    nombres_usados = set()
    for plan in aprendices_planes:
        datos_apr = plan["datos_aprendiz"]
        nombre_base = f"{datos_apr.get('nombre', '')} {datos_apr.get('apellidos', '')}".strip()
        nombre_hoja = _sanitizar_nombre_hoja(nombre_base)
        # Evitar duplicados
        original = nombre_hoja
        contador = 2
        while nombre_hoja in nombres_usados:
            sufijo = f" ({contador})"
            nombre_hoja = _sanitizar_nombre_hoja(original[:31 - len(sufijo)]) + sufijo
            contador += 1
        nombres_usados.add(nombre_hoja)

        ws = wb.create_sheet(title=nombre_hoja)
        _construir_hoja_aprendiz(ws, plan, datos_guia, instructor)

    wb.save(ruta_salida)
    return ruta_salida


def _construir_resumen(ws, aprendices_planes, datos_guia, instructor):
    # Encabezado
    ws["A1"] = "PORTAFOLIO — PLANES DE TRABAJO INDIVIDUAL"
    ws["A1"].font = Font(bold=True, size=14, color=DARK)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Programa: {datos_guia.get('programa', '')}"
    ws["A2"].font = Font(size=10, color=DARK)
    ws.merge_cells("A2:H2")

    ws["A3"] = f"Competencia: {datos_guia.get('competencia', '')}"
    ws["A3"].font = Font(size=10, color=DARK)
    ws.merge_cells("A3:H3")

    ws["A4"] = f"Instructor: {instructor.get('nombre', '')} · Fecha de emisión: {date.today().isoformat()}"
    ws["A4"].font = Font(size=10, italic=True, color=DARK)
    ws.merge_cells("A4:H4")

    # Fila en blanco
    fila_encabezados = 6

    # Encabezados de columnas
    encabezados = [
        "N°", "Nombre completo", "Correo", "Ficha", "Programa",
        "Correo enviado", "Fecha envío", "Ver detalle"
    ]
    for col, texto in enumerate(encabezados, 1):
        celda = ws.cell(row=fila_encabezados, column=col, value=texto)
        celda.font = Font(bold=True, color=WHITE)
        celda.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = _borde_gris()

    # Filas
    for i, plan in enumerate(aprendices_planes, start=1):
        datos_apr = plan["datos_aprendiz"]
        fila = fila_encabezados + i
        nombre_completo = f"{datos_apr.get('nombre', '')} {datos_apr.get('apellidos', '')}".strip()

        valores = [
            i,
            nombre_completo,
            datos_apr.get("correo", ""),
            str(datos_apr.get("ficha", "")),
            datos_apr.get("programa", ""),
            "✅ Sí" if plan.get("correo_enviado") else "⏳ Pendiente",
            plan.get("fecha_envio", ""),
            f'→ Ver hoja "{_sanitizar_nombre_hoja(nombre_completo)}"',
        ]
        for col, valor in enumerate(valores, 1):
            c = ws.cell(row=fila, column=col, value=valor)
            c.font = Font(size=9, color=DARK)
            c.border = _borde_gris()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i % 2 == 0:
                c.fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")

    # Anchos
    anchos = [5, 30, 30, 12, 25, 15, 15, 25]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _construir_hoja_aprendiz(ws, plan, datos_guia, instructor):
    datos_apr = plan["datos_aprendiz"]
    cronograma = plan.get("cronograma", [])
    nombre_completo = f"{datos_apr.get('nombre', '')} {datos_apr.get('apellidos', '')}".strip()

    # Encabezado
    ws["A1"] = "PLAN DE TRABAJO INDIVIDUAL"
    ws["A1"].font = Font(bold=True, size=14, color=DARK)
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Datos del aprendiz
    campos = [
        ("Nombre completo:", nombre_completo),
        ("Correo:", datos_apr.get("correo", "")),
        ("Ficha:", str(datos_apr.get("ficha", ""))),
        ("Programa:", datos_apr.get("programa", "") or datos_guia.get("programa", "")),
        ("Competencia:", datos_guia.get("competencia", "")),
        ("Proyecto formativo:", datos_guia.get("proyecto_formativo", "")),
        ("Fase:", datos_guia.get("fase_proyecto", "")),
    ]
    for i, (etiqueta, valor) in enumerate(campos, start=3):
        c_et = ws.cell(row=i, column=1, value=etiqueta)
        c_et.font = Font(bold=True, size=10, color=DARK)
        c_et.fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
        c_et.border = _borde_gris()
        c_val = ws.cell(row=i, column=2, value=valor)
        c_val.font = Font(size=10, color=DARK)
        c_val.border = _borde_gris()
        c_val.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)

    # Cronograma
    fila_titulo_crono = 3 + len(campos) + 1
    ws.cell(row=fila_titulo_crono, column=1, value="CRONOGRAMA DE ACTIVIDADES").font = Font(bold=True, size=12, color=DARK)
    ws.merge_cells(start_row=fila_titulo_crono, start_column=1, end_row=fila_titulo_crono, end_column=5)

    fila_headers = fila_titulo_crono + 1
    headers_crono = ["Actividad", "Descripción", "Inicio", "Entrega", "Entregado (Sí/No)"]
    for col, h in enumerate(headers_crono, 1):
        c = ws.cell(row=fila_headers, column=col, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borde_gris()

    for i, act in enumerate(cronograma, start=1):
        fila = fila_headers + i
        valores = [
            f"{act['titulo']} ({act['horas']:.0f} h)",
            act["descripcion"],
            act["fecha_inicio"].strftime("%d/%m/%Y") if hasattr(act["fecha_inicio"], "strftime") else str(act["fecha_inicio"]),
            act["fecha_entrega"].strftime("%d/%m/%Y") if hasattr(act["fecha_entrega"], "strftime") else str(act["fecha_entrega"]),
            "",  # el instructor lo llena manualmente
        ]
        for col, v in enumerate(valores, 1):
            c = ws.cell(row=fila, column=col, value=v)
            c.font = Font(size=9, color=DARK)
            c.border = _borde_gris()
            c.alignment = Alignment(vertical="center", wrap_text=True,
                                    horizontal="center" if col >= 3 else "left")
            if i % 2 == 0:
                c.fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")

    # Observaciones del instructor
    fila_obs = fila_headers + len(cronograma) + 2
    ws.cell(row=fila_obs, column=1, value="OBSERVACIONES DEL INSTRUCTOR").font = Font(bold=True, size=11, color=DARK)
    ws.merge_cells(start_row=fila_obs, start_column=1, end_row=fila_obs, end_column=5)

    ws.cell(row=fila_obs + 1, column=1, value="").fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
    ws.merge_cells(start_row=fila_obs + 1, start_column=1, end_row=fila_obs + 4, end_column=5)
    ws.cell(row=fila_obs + 1, column=1).border = _borde_gris()
    ws.cell(row=fila_obs + 1, column=1).alignment = Alignment(vertical="top", wrap_text=True)

    # Datos de envío
    fila_envio = fila_obs + 6
    envio_info = [
        ("Correo enviado a:", datos_apr.get("correo", "")),
        ("Estado:", "✅ Enviado" if plan.get("correo_enviado") else "⏳ Pendiente"),
        ("Fecha de envío:", plan.get("fecha_envio", "N/A")),
        ("PDF adjunto:", Path(plan.get("archivo_pdf", "")).name if plan.get("archivo_pdf") else ""),
    ]
    for i, (etq, val) in enumerate(envio_info):
        c_et = ws.cell(row=fila_envio + i, column=1, value=etq)
        c_et.font = Font(bold=True, size=9, color=DARK)
        c_val = ws.cell(row=fila_envio + i, column=2, value=val)
        c_val.font = Font(size=9, color=DARK)
        ws.merge_cells(start_row=fila_envio + i, start_column=2,
                       end_row=fila_envio + i, end_column=5)

    # Anchos
    anchos = [22, 40, 14, 14, 18]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
